"""
Performance Report Service

Queries JOBS + sessions + milestones from Firestore and builds a flat table
of job performance data with sub-task timing (computed from consecutive
milestone timestamps).

Flat table columns:
    Session Name, Job Start Time, Job Name, Question, Task, Time Taken,
    LLM Call, Dependency, Optimization, Job ID, Session ID, User ID

Admin-only — not user-facing.
"""
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone
from io import BytesIO
from v2.common.gcp import GcpManager
from v2.common.logger import add_log
from v2.modules.job_framework.manager.job.job_manager import JobManager
from v2.modules.session_framework.services.session.session_service import SessionService
import traceback


class PerformanceReportService:
    """
    Builds a flat performance report table from Firestore milestones.

    Each row = one milestone (sub-task) within a job.
    Duration = gap between consecutive milestone ``created_at`` timestamps.
    """

    def __init__(self):
        self._job_manager = JobManager()
        self._session_service = SessionService()
        try:
            gcp = GcpManager._get_instance()
            self._firestore = gcp._firestore_service
        except Exception:
            self._firestore = None

    # ── public API ─────────────────────────────────────────────────────

    async def get_performance_report(
        self,
        user_id: Optional[str] = None,
        session_id: Optional[str] = None,
        from_date: Optional[datetime] = None,
        to_date: Optional[datetime] = None,
    ) -> List[Dict[str, Any]]:
        """
        Build the flat performance table.

        Filters (all optional, combinable):
            user_id    – restrict to jobs owned by this user
            session_id – restrict to jobs in this session
            from_date  – jobs created on or after this datetime
            to_date    – jobs created on or before this datetime

        Returns a list of row dicts (one per milestone / sub-task).
        """
        try:
            jobs = await self._fetch_jobs(user_id, session_id, from_date, to_date)
            if not jobs:
                return []

            # Collect unique session IDs so we can batch-fetch labels
            session_ids = list({j.session_id for j in jobs if j.session_id})
            session_labels = await self._fetch_session_labels(session_ids)

            rows: List[Dict[str, Any]] = []
            for job in jobs:
                job_rows = await self._build_job_rows(job, session_labels)
                rows.extend(job_rows)

            return rows
        except Exception as e:
            add_log(f"PerformanceReportService: Error building report: {str(e)} | traceback: {traceback.format_exc()}")
            raise

    async def export_to_excel(
        self,
        user_id: Optional[str] = None,
        session_id: Optional[str] = None,
        from_date: Optional[datetime] = None,
        to_date: Optional[datetime] = None,
    ) -> BytesIO:
        """
        Generate an Excel (.xlsx) file from the flat performance table.

        Returns a BytesIO buffer ready for streaming as an HTTP response.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        rows = await self.get_performance_report(user_id, session_id, from_date, to_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Report"

        # ── header ──
        headers = [
            "Session Name", "Job Start Time", "Job Name", "Question",
            "Task", "Time Taken (s)", "LLM Call", "Dependency", "Optimization",
            "Job ID", "Session ID", "User ID",
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # ── data rows ──
        for row_idx, row in enumerate(rows, start=2):
            values = [
                row.get("session_name", ""),
                row.get("job_start_time", ""),
                row.get("job_name", ""),
                row.get("question", ""),
                row.get("task", ""),
                row.get("time_taken_seconds"),
                row.get("llm_call", ""),
                row.get("dependency", ""),
                row.get("optimization", ""),
                row.get("job_id", ""),
                row.get("session_id", ""),
                row.get("user_id", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx == 6 and val is not None:
                    cell.number_format = "0.0"
                    cell.alignment = Alignment(horizontal="right")

        # auto-width columns (cap at 50)
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── private helpers ────────────────────────────────────────────────

    async def _fetch_jobs(self, user_id, session_id, from_date, to_date):
        """Fetch jobs from Firestore with optional filters."""
        filters: List[tuple] = []

        if user_id:
            filters.append(("user_id", "==", user_id))
        if session_id:
            filters.append(("session_id", "==", session_id))
        if from_date:
            filters.append(("created_on", ">=", from_date))
        if to_date:
            filters.append(("created_on", "<=", to_date))

        # Only fetch analysis and qna job types (the ones with meaningful milestones)
        # We don't filter by job_type here to keep the report general —
        # jobs without milestones simply won't produce sub-task rows.

        if not self._firestore:
            add_log("PerformanceReportService: Firestore not available")
            return []

        try:
            from v2.modules.job_framework.models.job_model import JobResponse

            jobs_data = await self._firestore._query_collection(
                "JOBS",
                filters=filters if filters else None,
                order_by="created_on",
                order_direction="desc",
            )

            jobs = []
            for doc in jobs_data:
                try:
                    doc_id = doc.get("id", "")
                    jobs.append(self._doc_to_job_response(doc_id, doc))
                except Exception as conv_err:
                    add_log(f"PerformanceReportService: Skipping job doc {doc.get('id','?')}: {conv_err}")
            return jobs
        except Exception as e:
            add_log(f"PerformanceReportService: Error fetching jobs: {str(e)} | traceback: {traceback.format_exc()}")
            raise

    def _doc_to_job_response(self, doc_id: str, doc: Dict[str, Any]):
        """Convert a raw Firestore doc dict to a JobResponse-like object (lightweight)."""
        from v2.modules.job_framework.models.job_model import JobResponse, JobStatus

        status_raw = doc.get("status", "pending")
        try:
            status = JobStatus(status_raw)
        except ValueError:
            status = JobStatus.PENDING

        return JobResponse(
            job_id=doc_id,
            user_id=str(doc.get("user_id", "")),
            session_id=str(doc.get("session_id", "")),
            job_type=doc.get("job_type", ""),
            status=status,
            label=doc.get("label", ""),
            created_on=doc.get("created_on", datetime.now(timezone.utc)),
            started_at=doc.get("started_at"),
            completed_at=doc.get("completed_at"),
            error_message=doc.get("error_message"),
            error_type=doc.get("error_type"),
            execution_metadata=doc.get("execution_metadata"),
            result_metadata=doc.get("result_metadata"),
            job_config=doc.get("job_config"),
        )

    async def _fetch_session_labels(self, session_ids: List[str]) -> Dict[str, str]:
        """Return {session_id: label} for given session IDs."""
        labels: Dict[str, str] = {}
        for sid in session_ids:
            try:
                session = await self._session_service._get_session_by_id(sid)
                if session:
                    labels[sid] = session.label or ""
            except Exception:
                labels[sid] = ""
        return labels

    async def _build_job_rows(self, job, session_labels: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Build flat-table rows for a single job.

        Each milestone produces one row. Duration is the gap between
        consecutive milestones (first milestone's duration = gap from
        job.started_at to the milestone's created_at).
        """
        session_name = session_labels.get(job.session_id, "")
        job_start_time = _format_datetime(job.started_at or job.created_on)
        job_name = job.label or job.job_type or ""
        question = await self._get_job_question(job)

        # Fetch milestones (ordered by created_at ascending)
        milestones = await self._job_manager._get_job_milestones(job.job_id)

        if not milestones:
            # No milestones — emit a single summary row using job-level times
            duration = None
            if job.started_at and job.completed_at:
                duration = round((job.completed_at - job.started_at).total_seconds(), 1)
            return [
                {
                    "session_name": session_name,
                    "job_start_time": job_start_time,
                    "job_name": job_name,
                    "question": question,
                    "task": f"total ({job.status.value})",
                    "time_taken_seconds": duration,
                    "llm_call": "",
                    "dependency": "",
                    "optimization": "",
                    "job_id": job.job_id,
                    "session_id": job.session_id,
                    "user_id": job.user_id,
                }
            ]

        rows: List[Dict[str, Any]] = []
        prev_time = job.started_at or job.created_on

        for idx, ms in enumerate(milestones):
            task_name = _derive_task_name(ms)
            ms_time = ms.created_at

            # Compute duration between this milestone and the previous one
            duration = None
            if prev_time and ms_time:
                try:
                    # Ensure both are timezone-aware for subtraction
                    p = _ensure_aware(prev_time)
                    c = _ensure_aware(ms_time)
                    duration = round((c - p).total_seconds(), 1)
                    if duration < 0:
                        duration = 0.0
                except Exception:
                    duration = None

            # Derive dependency, optimization, llm_call from milestone.data
            ms_data = getattr(ms, "data", None) or {}
            dependency = ms_data.get("dependency", "sequential")
            optimization = "Consider async" if dependency == "parallelizable" else ""
            is_llm = ms_data.get("is_llm_call", False)
            llm_call = "Yes" if is_llm else "No"

            rows.append(
                {
                    # Repeat values for all rows (no blanks) so reports can process consistently
                    "session_name": session_name,
                    "job_start_time": job_start_time,
                    "job_name": job_name,
                    "question": question,
                    "task": task_name,
                    "time_taken_seconds": duration,
                    "llm_call": llm_call,
                    "dependency": dependency,
                    "optimization": optimization,
                    "job_id": job.job_id,
                    "session_id": job.session_id,
                    "user_id": job.user_id,
                }
            )
            prev_time = ms_time

        return rows

    async def _get_job_question(self, job) -> str:
        """
        Extract the question / query for a job.

        - Analysis jobs:  job_config["query"]
        - Simple QnA jobs: first query from JOBS/{chat_id}/Queries subcollection
        - Other jobs:     empty string
        """
        job_type = (job.job_type or "").lower()

        if job_type == "analysis":
            config = job.job_config or {}
            return config.get("query", "")

        if job_type == "qna":
            try:
                queries_path = f"JOBS/{job.job_id}/Queries"
                if self._firestore:
                    queries = await self._firestore._query_collection(
                        queries_path,
                        order_by="created_at",
                        order_direction="asc",
                        limit=1,
                    )
                    if queries:
                        return queries[0].get("query", "")
            except Exception:
                pass
            return ""

        return ""


# ── module-level helpers ───────────────────────────────────────────────

def _format_datetime(dt: Optional[datetime]) -> str:
    """Return ISO-ish human-readable string, or empty."""
    if dt is None:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(dt)


def _ensure_aware(dt: datetime) -> datetime:
    """Attach UTC tzinfo if naive."""
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _derive_task_name(milestone) -> str:
    """
    Derive a concise task / sub-phase name from a milestone.

    Priority: name → description → fallback.
    Strips common noise prefixes.
    """
    name = milestone.name or ""
    desc = milestone.description or ""

    # Prefer description if it's more informative than a machine-generated name
    if desc and not desc.startswith("milestone_"):
        raw = desc
    elif name and not name.startswith("milestone_"):
        raw = name
    elif desc:
        raw = desc
    else:
        raw = name or "unknown"

    # Clean up: strip prefixes like "milestone_<id>_"
    cleaned = raw
    if cleaned.startswith("milestone_"):
        # Format: milestone_<job_id>_<actual_name>
        parts = cleaned.split("_", 2)
        if len(parts) >= 3:
            cleaned = parts[2]

    # Truncate if very long
    if len(cleaned) > 80:
        cleaned = cleaned[:77] + "..."

    return cleaned
